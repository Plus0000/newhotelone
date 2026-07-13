"""
对每个 Excel 运维 sheet，比较"年度总价"列求和 vs TS 文件 subtotal 求和。
"""
import openpyxl
import os

EXCEL_DIR = '/Users/plus0/Desktop/技术交底-20260701/模块3：机电系统投资概算/设备材料安装运维表/'

# techId -> Excel 文件名 (根据 techName 对应)
EXCEL_MAP = {
    '1': '10、相变储热供暖技术.xlsx',           # 相变储热供暖技术
    '2': '1、IoT+数字孪生+AI前馈调节.xlsx',     # AI 边缘计算节能管控
    '3': '8、地源热泵_空气源热泵多能源耦合供热技术.xlsx',  # 多能互补热泵系统
    '4': '11、智能照明控制技术.xlsx',           # DALI 智能照明
    '5': '2、高效空调制冷机房技术.xlsx',        # 六管制机房节能
    '6': '4、冷水机组冷凝热回收技术.xlsx',      # 高效冷水机组
    '7': '5、冷却塔供冷技术.xlsx',              # 磁悬浮压缩机
    '8': '12、光储充一体化技术.xlsx',           # 光伏储能一体化
    '9': '3、洁净区域冷热源升级技术.xlsx',      # 能源计量监测
    '10': '6、蒸汽锅炉系统冷凝水余热回收技术.xlsx',  # 锅炉余热回收
    '11': '9、分时分区供暖技术.xlsx',           # 新风热回收
    '12': '7、蓄冷技术.xlsx',                   # 电梯能量回馈
}

# 运维表列: 选择|维护项目|维护分类|单位|数量|单价（元）|年度总价（万元）|年限|全维护周期总价|备注
# 年度总价在第 7 列
ANNUAL_COL = 7
# 数量在第 5 列, 单价（元）在第 6 列
QTY_COL = 5
PRICE_COL = 6

print('=== Excel 运维表 年度总价 vs 数量×单价 对比 ===\n')
print(f'{"techId":<8} {"年度总价和":<14} {"数量×单价和(万元)":<18} {"一致?":<6} {"异常行数":<8}')
print('-' * 60)

for tech_id, fname in EXCEL_MAP.items():
    path = os.path.join(EXCEL_DIR, fname)
    if not os.path.exists(path):
        print(f'{tech_id:<8} 文件不存在: {fname}')
        continue
    wb = openpyxl.load_workbook(path, data_only=True)
    if '运营维护' not in wb.sheetnames:
        print(f'{tech_id:<8} 无运营维护 sheet')
        continue
    ws = wb['运营维护']
    annual_sum = 0.0
    qty_price_sum = 0.0  # 数量 × 单价(元) / 10000 = 万元
    bad_rows = 0
    for r in range(2, ws.max_row + 1):
        qty = ws.cell(r, QTY_COL).value
        price = ws.cell(r, PRICE_COL).value  # 元
        annual = ws.cell(r, ANNUAL_COL).value  # 万元
        if annual is None or annual == '-' or annual == '':
            continue
        try:
            annual_val = float(annual)
        except (ValueError, TypeError):
            annual_val = 0
        annual_sum += annual_val

        # 数量 × 单价（万元）
        try:
            if qty is None or qty == '' or qty == '不限' or qty == '-':
                qty_val = 0
            else:
                qty_val = float(qty)
        except (ValueError, TypeError):
            qty_val = 0
        try:
            if price is None or price == '' or price == '-':
                price_val = 0
            else:
                price_val = float(price)
        except (ValueError, TypeError):
            price_val = 0
        subtotal = qty_val * price_val / 10000  # 转万元
        qty_price_sum += subtotal

        # 检查是否一致（允许 0.01 万元误差）
        if abs(subtotal - annual_val) > 0.01:
            bad_rows += 1
            print(f'  techId={tech_id} R{r}: 数量={qty} 单价={price}元 subtotal={subtotal:.4f}万 vs 年度总价={annual_val}万')

    ok = '✓' if abs(annual_sum - qty_price_sum) < 0.01 else '✗'
    print(f'{tech_id:<8} {annual_sum:<14.4f} {qty_price_sum:<18.4f} {ok:<6} {bad_rows:<8}')
