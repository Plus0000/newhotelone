#!/usr/bin/env python3
"""根据桌面 Excel 生成 src/data/standards.ts"""
import re
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
STANDARDS_TS = ROOT / "src" / "data" / "standards.ts"
EXCEL_PATH = Path("/Users/plus0/Desktop/技术交底-20260701/规范标准库(1)/规范标准全名录.xlsx")

CATEGORY_MAP = {
    "通用设计": "general_design",
    "设备及能效": "equipment_efficiency",
    "施工验收": "construction_acceptance",
    "运维能效管控": "operation_energy_control",
    "能源管理技术": "energy_management",
    "政策法规文件": "policy_regulation",
    "智能化": "intelligent",
    "工程技术": "engineering",
    "新能源/新型电网": "new_energy_grid",
    "绿色碳排": "green_carbon",
    "医院专属": "hospital_specific",
    "图集": "drawing_atlas",
    "能耗定额地标": "energy_quota_local",
}


def extract_existing_urls(text: str) -> dict[str, str]:
    """从现有 standards.ts 中提取 code -> url 映射。"""
    code_url: dict[str, str] = {}
    # 简单正则：捕获 code 后面最近一个 url: "..."
    entries = re.split(r"\{\s*id:", text)
    for entry in entries[1:]:
        code_match = re.search(r'code:\s*"([^"]+)"', entry)
        url_match = re.search(r'url:\s*"([^"]+)"', entry)
        if code_match and url_match:
            code_url[code_match.group(1)] = url_match.group(1)
    return code_url


_ATLAS_RE = re.compile(r"^(\d+[A-Z]\d+)\s+(.+)$")


def parse_standard(raw: str) -> tuple[str, str]:
    """解析《名称》（编号） -> (name, code)；兼容全角/半角括号、图集编号前缀。"""
    # 1. 名称（编号）
    m = re.match(r"《(.+?)》[（(](.+?)[）)]", raw)
    if m:
        return m.group(1), m.group(2)
    # 2. 图集类：编号 名称
    m2 = re.match(r"《(.+?)》", raw)
    if m2:
        body = m2.group(1).strip()
        atlas = _ATLAS_RE.match(body)
        if atlas:
            return atlas.group(2).strip(), atlas.group(1).strip()
        return body, ""
    # 3. 无书名号，尝试图集类
    atlas = _ATLAS_RE.match(raw.strip())
    if atlas:
        return atlas.group(2).strip(), atlas.group(1).strip()
    return raw, ""


def escape_string(s: str) -> str:
    return s.replace("\\", "\\\\").replace('"', '\\"')


def main() -> None:
    existing = STANDARDS_TS.read_text(encoding="utf-8")
    code_url = extract_existing_urls(existing)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    # 如果已存在 PDF 匹配映射，加入 pdfPath
    pdf_matches: dict[str, str] = {}
    pdf_matches_path = ROOT / "scripts" / "pdf-matches.json"
    if pdf_matches_path.exists():
        import json

        pdf_matches = json.loads(pdf_matches_path.read_text(encoding="utf-8"))

    entries: list[dict] = []
    current_cat: str | None = None
    for r in range(2, ws.max_row + 1):
        seq = ws.cell(r, 1).value
        cat = ws.cell(r, 2).value
        raw = ws.cell(r, 3).value
        location = ws.cell(r, 4).value
        entered = ws.cell(r, 5).value

        if cat:
            current_cat = cat
        if not current_cat:
            raise ValueError(f"第 {r} 行缺少分类")

        name, code = parse_standard(raw)
        category_key = CATEGORY_MAP.get(current_cat)
        if not category_key:
            raise ValueError(f"未映射分类: {current_cat}")

        entry = {
            "id": str(int(seq)),
            "category": category_key,
            "name": name,
            "code": code,
            "location": location,
            "hasPdf": bool(entered),
        }
        if code in code_url:
            entry["url"] = code_url[code]
        if entry["id"] in pdf_matches:
            entry["pdfPath"] = f"{entry['id']}.pdf"
        entries.append(entry)

    lines = [
        "// 规范标准库 — 预置参考数据",
        "// 汇集医院建筑国/地/行标，为节能方案制定、节能计算、用户问答提供依据",
        f"// 数据来源：规范标准全名录.xlsx（{len(entries)}条）",
        "",
        "export type StandardCategory =",
    ]
    for cat in sorted(set(CATEGORY_MAP.values())):
        lines.append(f"  | '{cat}'")
    lines.extend([
        ";",
        "",
        "export interface StandardEntry {",
        "  id: string;",
        "  category: StandardCategory;",
        "  name: string;",
        "  code: string;",
        "  /** 原文链接 */",
        "  url?: string;",
        "  /** PDF 在 Supabase Storage 中的路径 */",
        "  pdfPath?: string;",
        "}",
        "",
        "export const standards: StandardEntry[] = [",
    ])

    for e in entries:
        lines.append("  {")
        lines.append(f'    id: "{e["id"]}",')
        lines.append(f'    category: "{e["category"]}",')
        lines.append(f'    name: "{escape_string(e["name"])}",')
        lines.append(f'    code: "{escape_string(e["code"])}",')
        if "url" in e:
            lines.append(f'    url: "{e["url"]}",')
        if "pdfPath" in e:
            lines.append(f'    pdfPath: "{e["pdfPath"]}",')
        lines.append("  },")

    lines.extend([
        "];",
        "",
    ])

    STANDARDS_TS.write_text("\n".join(lines), encoding="utf-8")
    print(f"已生成 {len(entries)} 条标准到 {STANDARDS_TS}")

    # 同时输出 JSON 供上传脚本读取
    import json

    json_path = ROOT / "scripts" / "standards.json"
    json_path.write_text(
        json.dumps(entries, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"已输出 JSON 到 {json_path}")


if __name__ == "__main__":
    main()
