#!/usr/bin/env python3
"""读取 12 个设备表 Excel，输出为 TS techDefaultInvestments 数据。

每个 Excel 4 个 sheet：
- 设备表：name/specification/unit/quantity/unitPrice/isMainEquipment/powerKw/remark
- 材料表：name/specification/unit/quantity/unitPrice/remark
- 安装调试：name/unit/quantity/unitPrice/remark
- 运营维护：name/unit/quantity/unitPrice/remark (+ maintenanceCategory 字段)
"""

import json
import re
from pathlib import Path

from openpyxl import load_workbook

DEVICE_DIR = Path("/Users/plus0/Desktop/技术交底-20260701/模块3：机电系统投资概算/设备材料安装运维表")

# Excel 文件名 -> techId 映射（按 techEntries 顺序）
FILE_TO_TECH_ID = {
    "10、相变储热供暖技术.xlsx": "1",
    "1、IoT+数字孪生+AI前馈调节.xlsx": "2",
    "8、地源热泵_空气源热泵多能源耦合供热技术.xlsx": "3",
    "11、智能照明控制技术.xlsx": "4",
    "3、洁净区域冷热源升级技术.xlsx": "5",
    "2、高效空调制冷机房技术.xlsx": "6",
    "4、冷水机组冷凝热回收技术.xlsx": "7",
    "5、冷却塔供冷技术.xlsx": "8",
    "6、蒸汽锅炉系统冷凝水余热回收技术.xlsx": "9",
    "7、蓄冷技术.xlsx": "10",
    "9、分时分区供暖技术.xlsx": "11",
    "12、光储充一体化技术.xlsx": "12",
}


def parse_bool(v):
    if v is None:
        return False
    s = str(v).strip()
    return s in ("☑", "√", "是", "Y", "y", "true", "True", "1")


def parse_num(v):
    if v is None or v == "":
        return 0
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0


def read_equipment(sheet):
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        rows.append({
            "name": str(row[1]).strip(),
            "category": str(row[2] or "").strip(),
            "specification": str(row[3] or "").strip(),
            "unit": str(row[4] or "").strip(),
            "quantity": parse_num(row[5]),
            "unitPrice": parse_num(row[6]),
            "isMainEquipment": parse_bool(row[9]) if len(row) > 9 else False,
            "powerKw": parse_num(row[10]) if len(row) > 10 else 0,
            "remark": str(row[8] or "").strip() if len(row) > 8 else "",
        })
    return rows


def read_materials(sheet):
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        rows.append({
            "name": str(row[1]).strip(),
            "category": str(row[2] or "").strip(),
            "specification": str(row[3] or "").strip(),
            "unit": str(row[4] or "").strip(),
            "quantity": parse_num(row[5]),
            "unitPrice": parse_num(row[6]),
            "remark": str(row[8] or "").strip() if len(row) > 8 else "",
        })
    return rows


def read_installation(sheet):
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        rows.append({
            "name": str(row[1]).strip(),
            "category": str(row[2] or "").strip(),
            "specification": "",
            "unit": str(row[3] or "").strip(),
            "quantity": parse_num(row[4]),
            "unitPrice": parse_num(row[5]),
            "remark": str(row[7] or "").strip() if len(row) > 7 else "",
        })
    return rows


def read_maintenance(sheet):
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        category = str(row[2] or "").strip() if len(row) > 2 else ""
        rows.append({
            "name": str(row[1]).strip(),
            "category": category,
            "specification": "",
            "unit": str(row[3] or "").strip(),
            "quantity": parse_num(row[4]),
            "unitPrice": parse_num(row[5]),
            "remark": str(row[9] or "").strip() if len(row) > 9 else "",
            "maintenanceCategory": category,
            "maintenanceYears": parse_num(row[7]) if len(row) > 7 else 0,
            "totalLifecycleCost": parse_num(row[8]) if len(row) > 8 else 0,
        })
    return rows


def main():
    all_data = {}
    for filename, tech_id in FILE_TO_TECH_ID.items():
        path = DEVICE_DIR / filename
        if not path.exists():
            print(f"⚠️  Not found: {path}")
            continue
        wb = load_workbook(path, data_only=True)
        all_data[tech_id] = {
            "equipment": read_equipment(wb["设备表"]),
            "materials": read_materials(wb["材料表"]),
            "installation": read_installation(wb["安装调试"]),
            "maintenance": read_maintenance(wb["运营维护"]),
        }
        print(f"✓ {filename} (techId={tech_id}): "
              f"equip={len(all_data[tech_id]['equipment'])}, "
              f"mat={len(all_data[tech_id]['materials'])}, "
              f"inst={len(all_data[tech_id]['installation'])}, "
              f"maint={len(all_data[tech_id]['maintenance'])}")

    # 输出为 JSON
    out_path = Path("/tmp/tech_investments.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    print(f"\n→ JSON saved to {out_path}")


if __name__ == "__main__":
    main()
