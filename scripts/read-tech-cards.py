#!/usr/bin/env python3
"""读取技术卡片 Excel 完整数据，输出为 JSON 供后续录入 TS。"""

import json
from pathlib import Path

from openpyxl import load_workbook

EXCEL = Path("/Users/plus0/Desktop/技术交底-20260701/模块2：节能方案筛选/节能技术全量信息（技术卡片）.xlsx")


def main():
    wb = load_workbook(EXCEL, data_only=True)
    sheet = wb["节能技术全量信息（技术卡片）"]

    # 列结构（从探测报告）:
    # A:序号 B:节能技术名称 C:作用系统 D:能耗种类 E:技术互斥
    # F:基准节能率区间 G-J:节能率取值1-4 K:核心依据说明
    # L:分类 M:技术原理 N:解决运营痛点
    # O-S:适用边界条件（医院类型/最小面积/气候分区/能源系统/适用科室）
    # T:固定投资额指标 U:年运行能耗 V:静态投资回报期
    rows = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        if not row[0] or not row[1]:
            continue
        rows.append({
            "序号": row[0],
            "节能技术名称": row[1],
            "作用系统": row[2],
            "能耗种类": row[3],
            "技术互斥": row[4],
            "基准节能率区间": row[5],
            "取值1": row[6],
            "取值2": row[7],
            "取值3": row[8],
            "取值4": row[9],
            "核心依据说明": row[10],
            "分类": row[11],
            "技术原理": row[12],
            "解决运营痛点": row[13],
            "医院类型及规模": row[14],
            "最小建筑面积": row[15],
            "气候分区": row[16],
            "能源系统类型": row[17],
            "适用科室": row[18],
            "固定投资额指标": row[19],
            "年运行能耗": row[20],
            "静态投资回报期": row[21],
        })

    print(f"共 {len(rows)} 条技术")
    print(json.dumps(rows, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
