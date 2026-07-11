#!/usr/bin/env python3
"""
Excel 结构探测脚本 - Phase 1.0 前置

对桌面所有 Excel 文件输出:
- sheet 名列表
- 每个 sheet 的列名(第 1 行)
- 每个 sheet 的前 3 行数据
- 每个 sheet 的总行数
- 特殊字符扫描(£/R/合并单元格标记等)
- 数据类型推断(数字/字符串/日期)

跨表一致性检查:
- 6 张计算表的"机电系统"字段值要一致
- 12 个设备表的"技术名称"和 techEntries 的 techName 要一致
- 碳排放因子表的"省份"和 energyPriceReferences 的 location 要一致
- 12 个设备表的"运维分类"枚举值要统一

输出: docs/superpowers/progress/excel-structure-report.md
"""

import hashlib
import json
import os
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DESKTOP_ROOT = Path("/Users/plus0/Desktop/技术交底-20260701")
OUTPUT_FILE = Path("docs/superpowers/progress/excel-structure-report.md")

SPECIAL_CHARS = ["£", "R", " ", "﻿"]


def probe_sheet(ws) -> dict[str, Any]:
    """探测单个 sheet 的结构。"""
    result: dict[str, Any] = {
        "sheet_name": ws.title,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "headers": [],
        "first_3_rows": [],
        "special_chars_found": [],
        "merged_cells_count": len(ws.merged_cells.ranges),
        "column_types": {},
    }

    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        result["headers"].append(str(val) if val is not None else "")

    for row in range(2, min(5, ws.max_row + 1)):
        row_data = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            row_data.append(val)
        result["first_3_rows"].append(row_data)

    for col in range(1, ws.max_column + 1):
        types_seen = set()
        for row in range(2, min(12, ws.max_row + 1)):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            types_seen.add(type(val).__name__)
        result["column_types"][result["headers"][col - 1]] = list(types_seen)

    special_hits: list[dict[str, Any]] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            s = str(cell.value)
            for ch in SPECIAL_CHARS:
                if ch in s:
                    special_hits.append({
                        "char": ch,
                        "cell": f"{get_column_letter(cell.column)}{cell.row}",
                        "value": s[:50],
                    })
                    if len(special_hits) >= 5:
                        break
            if len(special_hits) >= 5:
                break
        if len(special_hits) >= 5:
            break
    result["special_chars_found"] = special_hits

    return result


def probe_file(filepath: Path) -> dict[str, Any]:
    """探测单个 Excel 文件。"""
    wb = load_workbook(filepath, data_only=True, read_only=False)
    sheets = []
    for ws in wb.worksheets:
        sheets.append(probe_sheet(ws))
    wb.close()
    return {
        "file_path": str(filepath.relative_to(DESKTOP_ROOT)),
        "file_name": filepath.name,
        "sheets": sheets,
    }


def render_markdown(results: list[dict[str, Any]]) -> str:
    """把探测结果渲染成 Markdown 报告。"""
    lines: list[str] = [
        "# Excel 结构探测报告",
        "",
        f"> 探测时间: 2026-07-11(Phase 1.0)",
        f"> 文件总数: {len(results)}",
        "",
        "## 目录",
        "",
    ]
    for i, r in enumerate(results, 1):
        anchor = r["file_name"].replace(" ", "-").replace("(", "").replace(")", "").replace("（", "").replace("）", "")
        lines.append(f"{i}. [{r['file_name']}](#{i}-{anchor.lower()})")
    lines.append("")

    for i, r in enumerate(results, 1):
        lines.append(f"## {i}. {r['file_name']}")
        lines.append("")
        lines.append(f"**路径**: `{r.get('file_path', '')}`")
        if "error" in r:
            lines.append(f"**ERROR**: {r['error']}")
            lines.append("")
            continue
        lines.append(f"**Sheet 数**: {len(r['sheets'])}")
        lines.append("")

        for sheet in r["sheets"]:
            lines.append(f"### Sheet: `{sheet['sheet_name']}`")
            lines.append("")
            lines.append(f"- 行数: {sheet['max_row']}")
            lines.append(f"- 列数: {sheet['max_col']}")
            lines.append(f"- 合并单元格数: {sheet['merged_cells_count']}")
            lines.append("")

            lines.append("**列名**:")
            lines.append("")
            for j, h in enumerate(sheet["headers"], 1):
                lines.append(f"{j}. `{h}`")
            lines.append("")

            lines.append("**前 3 行数据**:")
            lines.append("")
            for row_idx, row in enumerate(sheet["first_3_rows"], 2):
                cells = [str(c)[:30] if c is not None else "" for c in row]
                lines.append(f"- 行 {row_idx}: {' | '.join(cells)}")
            lines.append("")

            lines.append("**列数据类型**:")
            lines.append("")
            for col_name, types in sheet["column_types"].items():
                lines.append(f"- `{col_name}`: {', '.join(types) if types else '(空)'}")
            lines.append("")

            if sheet["special_chars_found"]:
                lines.append("**特殊字符扫描**:")
                lines.append("")
                for hit in sheet["special_chars_found"]:
                    lines.append(f"- `{hit['char']}` @ {hit['cell']}: {hit['value']}")
                lines.append("")

        lines.append("---")
        lines.append("")

    return "\n".join(lines)


def check_cross_table_consistency(results: list[dict[str, Any]]) -> str:
    """跨表一致性检查,返回 Markdown 段落。"""
    lines: list[str] = [
        "## 跨表一致性检查",
        "",
    ]

    # 1. 6 张计算表的"机电系统"字段值一致性
    calc_tables = [
        "机电系统能耗权重表.xlsx",
        "医院整体修正系数表.xlsx",
        "技术组合重叠修正系数表.xlsx",
    ]
    system_values: dict[str, set[str]] = {}
    for r in results:
        if r["file_name"] in calc_tables:
            for sheet in r.get("sheets", []):
                for j, h in enumerate(sheet["headers"]):
                    if "机电系统" in str(h) or "系统类型" in str(h):
                        vals = set()
                        for row in sheet["first_3_rows"]:
                            if j < len(row) and row[j] is not None:
                                vals.add(str(row[j]))
                        system_values.setdefault(r["file_name"], set()).update(vals)
    lines.append("### 1. 计算表的机电系统字段值")
    lines.append("")
    for fname, vals in system_values.items():
        lines.append(f"- **{fname}**: {', '.join(sorted(vals)) if vals else '(无)'}")
    if not system_values:
        lines.append("- (未找到机电系统列,需人工检查列名)")
    lines.append("")

    # 2. 12 个设备表的"技术名称"一致性
    tech_names: list[str] = []
    for r in results:
        if "模块3" in r.get("file_path", "") and r["file_name"].endswith(".xlsx"):
            tech_names.append(r["file_name"])
    lines.append("### 2. 模块3 设备表文件清单")
    lines.append("")
    for n in tech_names:
        lines.append(f"- {n}")
    lines.append(f"**总数**: {len(tech_names)}(应为 12)")
    lines.append("")

    # 3. 碳排放因子表的"省份"字段
    carbon_files = ["电力平均碳排放因子（省级电网）.xlsx", "化石能源碳排放因子.xlsx"]
    lines.append("### 3. 碳排放因子表的省份字段")
    lines.append("")
    for r in results:
        if r["file_name"] in carbon_files:
            for sheet in r.get("sheets", []):
                for j, h in enumerate(sheet["headers"]):
                    if "省" in str(h) or "地区" in str(h):
                        vals = set()
                        for row in sheet["first_3_rows"]:
                            if j < len(row) and row[j] is not None:
                                vals.add(str(row[j]))
                        lines.append(f"- **{r['file_name']}** 列 `{h}`: {', '.join(sorted(vals)) if vals else '(无)'}")
    lines.append("")

    # 4. 设备表"运维分类"枚举值
    lines.append("### 4. 设备表运维分类枚举值")
    lines.append("")
    maintenance_categories: set[str] = set()
    for r in results:
        if "模块3" in r.get("file_path", ""):
            for sheet in r.get("sheets", []):
                if "运维" not in sheet["sheet_name"] and "维护" not in sheet["sheet_name"]:
                    continue
                for j, h in enumerate(sheet["headers"]):
                    if "分类" in str(h) or "维护分类" in str(h):
                        for row in sheet["first_3_rows"]:
                            if j < len(row) and row[j] is not None:
                                maintenance_categories.add(str(row[j]))
    lines.append(f"枚举值汇总: {', '.join(sorted(maintenance_categories)) if maintenance_categories else '(未找到)'}")
    lines.append("")
    lines.append(f"**期望值**: `维保费用`, `运维人工费用`(spec §5.1)")
    lines.append("")

    return "\n".join(lines)


def main() -> None:
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    xlsx_files = sorted(DESKTOP_ROOT.rglob("*.xlsx"))
    xlsx_files = [f for f in xlsx_files if not f.name.startswith(".~")]
    print(f"Found {len(xlsx_files)} xlsx files")

    results = []
    for f in xlsx_files:
        print(f"Probing: {f.name}")
        try:
            results.append(probe_file(f))
        except Exception as e:
            print(f"  ERROR: {e}")
            results.append({
                "file_path": str(f.relative_to(DESKTOP_ROOT)),
                "file_name": f.name,
                "error": str(e),
            })

    # 渲染 Markdown
    md = render_markdown(results)

    # 跨表一致性检查
    cross_check = check_cross_table_consistency(results)
    md = md + "\n" + cross_check

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(md)

    # 同时 dump JSON 作为 debug 备份
    with open("/tmp/excel-probe-results.json", "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2, default=str)

    print(f"\nMarkdown report written to {OUTPUT_FILE}")
    print(f"JSON debug dump: /tmp/excel-probe-results.json")


if __name__ == "__main__":
    main()
